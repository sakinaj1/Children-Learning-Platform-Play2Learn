from flask import Flask, request, jsonify, send_from_directory, session
import pandas as pd
import os
import time
from openpyxl import load_workbook

app = Flask(__name__)
app.secret_key = "secret123"

FILE = "excel.xlsx"

EXCEL_FILE = r'D:\Projects\HCI\HCI-FINAL\final\excel.xlsx'
COLUMNS = ['Name', 'Email', 'Password', 'Status', 'ChildName', 'Age', 'Interests', 'Goals', 'ChildID', 'TotalScore', 'TotalStars', 'GamesPlayed']

# init_user route removed, handled by signup/questionnaire

# ─── Excel Functions ─────────────────────────

def load_df():
    if os.path.exists(EXCEL_FILE):
        try:
            df = pd.read_excel(EXCEL_FILE, sheet_name=0)
            # Ensure all columns exist
            for col in COLUMNS:
                if col not in df.columns:
                    df[col] = ''
            return df.fillna('')
        except Exception:
            return pd.DataFrame(columns=COLUMNS)
    else:
        df = pd.DataFrame(columns=COLUMNS)
        df.to_excel(EXCEL_FILE, index=False, sheet_name='Users')
        return df

def save_df(df):
    if not os.path.exists(EXCEL_FILE):
        df.to_excel(EXCEL_FILE, index=False, sheet_name='Users')
    else:
        wb = load_workbook(EXCEL_FILE)
        first_sheet = wb.sheetnames[0]
        wb.close()
        
        with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=first_sheet, index=False)

# ─── Serve HTML ─────────────────────────────

@app.route('/')
def home():
    return send_from_directory('.', 'index.html')

@app.route('/<path:filename>')
def files(filename):
    if filename.endswith(".html") and filename not in ["index.html", "questionnaire.html"]:
        filepath = os.path.join('.', filename)
        if os.path.exists(filepath):
            with open(filepath, 'r', encoding='utf-8') as f:
                content = f.read()
            script = '''
<script>
    setInterval(function() {
        let limit = localStorage.getItem('dailyTimeLimit');
        let isHardStop = localStorage.getItem('hardStopEnabled');
        let start = localStorage.getItem('sessionStartTime');
        if (limit && isHardStop === 'true' && start) {
            let minutesPassed = (Date.now() - parseInt(start)) / 60000;
            if (minutesPassed >= parseInt(limit)) {
                localStorage.removeItem('sessionStartTime'); 
                alert("⏳ Daily Screen Time Limit Reached! See you tomorrow!");
                window.location.href = "index.html";
            }
        }
    }, 5000);
</script>
</body>
'''
            if '</body>' in content:
                content = content.replace('</body>', script)
                return content
    return send_from_directory('.', filename)

# ─── SIGNUP ─────────────────────────────────

@app.route('/api/signup', methods=['POST'])
def signup():
    data = request.get_json()

    name = data.get('name').strip()
    email = data.get('email').strip().lower()
    password = data.get('password').strip()

    df = load_df()

    if email in df['Email'].astype(str).str.lower().values:
        return jsonify({'success': False, 'message': 'Email already exists'})

    new_row = {
        'Name': name,
        'Email': email,
        'Password': password,
        'Status': 'pending',
        'ChildName': '',
        'Age': '',
        'Interests': '',
        'Goals': '',
        'ChildID': '',
        'TotalScore': 0,
        'TotalStars': 0,
        'GamesPlayed': 0
    }

    df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
    save_df(df)

    session['email'] = email

    return jsonify({'success': True, 'redirect': 'questionnaire.html'})

# ─── LOGIN ─────────────────────────────────

@app.route('/api/login', methods=['POST'])
def login():
    data = request.get_json()

    email = data.get('email').strip().lower()
    password = data.get('password').strip()

    df = load_df()

    df['Email'] = df['Email'].astype(str).str.strip().str.lower()
    df['Password'] = df['Password'].astype(str).str.strip()

    user = df[df['Email'] == email]

    if user.empty:
        return jsonify({'success': False, 'message': 'User not found'})

    user = user.iloc[0]

    if user['Password'] != password:
        return jsonify({'success': False, 'message': 'Wrong password'})

    session['email'] = email
    
    if user['Status'] == 'completed':
        session['child_id'] = user['ChildID']
        return jsonify({'success': True, 'redirect': 'home.html'})
    else:
        return jsonify({'success': True, 'redirect': 'questionnaire.html'})

# ─── QUESTIONNAIRE ─────────────────────────
@app.route('/api/questionnaire', methods=['POST'])
def questionnaire():
    data = request.get_json()
    email = session.get('email')

    child_name = data.get('childName')
    age = data.get('age')
    interests = ", ".join(data.get('interests', []))
    goals = ", ".join(data.get('goals', []))

    df = load_df()

    mask = df['Email'].astype(str).str.lower() == email

    if not mask.any():
        return jsonify({'success': False, 'message': 'User not found'})

    # 🔥 CREATE UNIQUE CHILD ID
    child_id = "CH" + str(int(time.time()))

    df.loc[mask, 'ChildName'] = child_name
    df.loc[mask, 'Age'] = age
    df.loc[mask, 'Interests'] = interests
    df.loc[mask, 'Goals'] = goals
    df.loc[mask, 'Status'] = 'completed'

    # 🔥 SAVE CHILD ID
    df.loc[mask, 'ChildID'] = child_id

    save_df(df)

    # 🔥 STORE IN SESSION
    session['child_id'] = child_id

    return jsonify({'success': True, 'redirect': 'home.html'})


@app.route('/get_child_id')
def get_child_id():
    return jsonify({
        "child_id": session.get("child_id")
    })
# ─── me ─────────────────────────
@app.route('/api/me', methods=['GET'])
def me():
    email = session.get('email')

    if not email:
        return jsonify({'loggedIn': False})

    df = load_df()

    # clean data
    df['Email'] = df['Email'].astype(str).str.strip().str.lower()

    email = email.strip().lower()

    user = df[df['Email'] == email]

    if user.empty:
        return jsonify({'loggedIn': False})

    user = user.iloc[0]

    return jsonify({
        'loggedIn': True,
        'name': user['Name'],
        'childName': user['ChildName'],
        'age': user['Age'],
        'status': user['Status'],
        'stars': int(user['TotalStars'] or 0),
        'gamesPlayed': int(user['GamesPlayed'] or 0)
    })

@app.route('/api/update_profile', methods=['POST'])
def update_profile():
    data = request.json
    email = session.get('email')
    if not email:
        return jsonify({"success": False, "message": "Not logged in"}), 401

    df = load_df()
    mask = df['Email'].astype(str).str.strip().str.lower() == email.strip().lower()

    if mask.any():
        if 'childName' in data:
            df.loc[mask, 'ChildName'] = data['childName']
        if 'age' in data:
            df.loc[mask, 'Age'] = data['age']
        save_df(df)
        return jsonify({"success": True})
    
    return jsonify({"success": False, "message": "User not found"})

@app.route('/api/verify_password', methods=['POST'])
def verify_password():
    data = request.json
    email = session.get('email')
    if not email:
        return jsonify({"success": False, "message": "Not logged in"}), 401

    provided_password = str(data.get('password', '')).strip()
    df = load_df()
    mask = df['Email'].astype(str).str.strip().str.lower() == email.strip().lower()
    
    if not mask.any():
        return jsonify({"success": False, "message": "User not found"}), 404

    user_row = df[mask].iloc[0]
    actual_password = str(user_row['Password']).strip()
    
    if provided_password == actual_password:
        return jsonify({"success": True})
    else:
        return jsonify({"success": False, "message": "Incorrect password"})

# ─── LOCK/UNLOCK + RESTRICTION ─────────────────────────
@app.route("/can_play_game")
def can_play_game():
    game = request.args.get("game")
    child_id = session.get("child_id")   # 🔥 FIX

    wb = load_workbook(FILE)

    if "restrictions" not in wb.sheetnames:
        return jsonify({"status": "allowed"})

    sheet = wb["restrictions"]

    difficulty = "easy"
    blocked = []

    for row in sheet.iter_rows(values_only=True):
        if row[0] == child_id:
            difficulty = row[1]
            blocked = row[2].split(",") if row[2] else []

    game_levels = {
        "alphabet": "easy",
        "numbers": "easy",
        "colors": "easy",
        "shapes": "easy",
        "memory": "medium",
        "math": "medium",
        "spelling": "hard",
        "puzzle": "hard"
    }

    if game in blocked:
        return jsonify({"status": "locked"})

    if difficulty == "easy" and game_levels[game] != "easy":
        return jsonify({"status": "locked"})

    if difficulty == "medium" and game_levels[game] == "hard":
        return jsonify({"status": "locked"})

    return jsonify({"status": "allowed"})


#restriction:
@app.route("/set_restrictions", methods=["POST"])
def set_restrictions():
    data = request.json

    child_id = session.get("child_id")
    difficulty = data.get("difficulty")
    blocked_games = ",".join(data.get("blockedGames", []))

    wb = load_workbook(FILE)

    if "restrictions" not in wb.sheetnames:
        sheet = wb.create_sheet("restrictions")
    else:
        sheet = wb["restrictions"]

    # REMOVE OLD ENTRY
    for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if row[0] == child_id:
            sheet.delete_rows(i)
            break

    # SAVE NEW
    sheet.append([child_id, difficulty, blocked_games])

    wb.save(FILE)

    return jsonify({"success": True})

# Score Submission

@app.route("/submit_score", methods=["POST"])
def submit_score():
    data = request.json
    score = data.get("score")
    stars = data.get("stars")
    game = data.get("game")

    email = session.get("email")
    if not email:
        return jsonify({"message": "Not logged in"}), 401

    df = load_df()
    mask = df['Email'].astype(str).str.lower() == email.lower()

    if mask.any():
        current_score = int(df.loc[mask, 'TotalScore'].values[0] or 0)
        current_stars = int(df.loc[mask, 'TotalStars'].values[0] or 0)
        current_games = int(df.loc[mask, 'GamesPlayed'].values[0] or 0)

        df.loc[mask, 'TotalScore'] = current_score + score
        df.loc[mask, 'TotalStars'] = current_stars + stars
        df.loc[mask, 'GamesPlayed'] = current_games + 1

        save_df(df)
        
        # Log activity
        if game:
            try:
                wb = load_workbook(EXCEL_FILE)
                if "activity_log" not in wb.sheetnames:
                    sheet = wb.create_sheet("activity_log")
                    sheet.append(["Email", "Game", "Date", "Score", "Stars", "TimeSpent"])
                else:
                    sheet = wb["activity_log"]
                
                time_spent = data.get("time_spent", 0)
                today = pd.Timestamp.now().strftime("%Y-%m-%d")
                sheet.append([email.lower(), game, today, score, stars, time_spent])
                wb.save(EXCEL_FILE)
            except Exception as e:
                print(f"Error saving activity log: {e}")

    return jsonify({"message": "Score updated"})

# Reward Purchase
@app.route("/api/buy_reward", methods=["POST"])
def buy_reward():
    data = request.json
    cost = data.get("cost")
    
    email = session.get("email")
    if not email:
        return jsonify({"success": False, "message": "Not logged in"}), 401

    df = load_df()
    mask = df['Email'].astype(str).str.lower() == email.lower()

    if mask.any():
        current_stars = int(df.loc[mask, 'TotalStars'].values[0] or 0)
        
        if current_stars >= cost:
            df.loc[mask, 'TotalStars'] = current_stars - cost
            save_df(df)
            return jsonify({"success": True, "stars": current_stars - cost})
        else:
            return jsonify({"success": False, "message": "Not enough stars!"})
            
    return jsonify({"success": False, "message": "User not found"})

# Progress 
@app.route("/get_progress")
def get_progress():
    email = session.get("email")
    if not email:
        return jsonify({})

    df = load_df()
    mask = df['Email'].astype(str).str.lower() == email.lower()

    if mask.any():
        row = df[mask].iloc[0]
        games_played = int(row['GamesPlayed'] or 0)
        progress = (games_played / 5) * 100
        
        # Calculate dynamic metrics based on activity_log
        weekly_stars = [0, 0, 0, 0, 0, 0, 0] # Mon to Sun
        skill_times = {"alphabet": 0, "numbers": 0, "shapes": 0, "colors": 0}
        total_time_spent = 0
        game_times = {} # To track top 3 games
        played_dates = set()
        played_today = set()

        
        try:
            wb = load_workbook(EXCEL_FILE)
            if "activity_log" in wb.sheetnames:
                sheet = wb["activity_log"]
                today = pd.Timestamp.now()
                current_week = today.isocalendar().week
                current_year = today.isocalendar().year
                
                for r in sheet.iter_rows(min_row=2, values_only=True):
                    if r[0] and str(r[0]).lower() == email.lower():
                        game_name = str(r[1]).lower() if r[1] else ""
                        date_val = str(r[2])
                        row_stars = 0
                        
                        try:
                            row_stars = int(r[4] or 0) # stars column
                        except: pass

                        time_val = 0
                        try:
                            if len(r) > 5 and r[5]:
                                time_val = int(r[5] or 0)
                                total_time_spent += time_val
                        except: pass

                        try:
                            # parse date string like YYYY-MM-DD
                            row_date_ds = pd.to_datetime(date_val)
                            row_date = row_date_ds.date()
                            played_dates.add(row_date)
                            if row_date == today.date() and game_name:
                                played_today.add(game_name)
                                
                            # Refresh out after each week: check if same calendar week
                            if row_date_ds.isocalendar().week == current_week and row_date_ds.isocalendar().year == current_year:
                                weekday = row_date_ds.weekday() # 0=Mon, 6=Sun
                                weekly_stars[weekday] += row_stars
                        except:
                            pass
                            
                        # Skill tracking by % of time given
                        if game_name in skill_times:
                            skill_times[game_name] += time_val
                            
                        if game_name:
                            game_times[game_name] = game_times.get(game_name, 0) + time_val
                            
        except Exception as e:
            print(f"Error reading activity log: {e}")
            
        # Convert skill times to % of total time
        skills = {}
        for s in ["alphabet", "numbers", "shapes", "colors"]:
            if total_time_spent > 0:
                skills[s] = int((skill_times[s] / total_time_spent) * 100)
            else:
                skills[s] = 0
        
        achievements = []
        # Recent achievements show top 3 games
        sorted_games = sorted(game_times.items(), key=lambda x: x[1], reverse=True)
        top_games = [g[0].capitalize() for g in sorted_games[:3]]
        
        if top_games:
            medals = ["🥇", "🥈", "🥉"]
            for idx, g in enumerate(top_games):
                achievements.append(f"{medals[idx]} Top Game: {g}")
        else:
            achievements.append("🌱 Play games to rank!")

        # Calculate streak
        played_dates_list = sorted(list(played_dates), reverse=True)
        streak = 0
        current_check = today.date()
        if played_dates_list:
            if played_dates_list[0] == current_check or played_dates_list[0] == current_check - pd.Timedelta(days=1):
                current_check = played_dates_list[0]
                for d in played_dates_list:
                    if d == current_check:
                        streak += 1
                        current_check -= pd.Timedelta(days=1)
                    else:
                        break

        return jsonify({
            "score": int(row['TotalScore'] or 0),
            "stars": int(row['TotalStars'] or 0),
            "progress": progress,
            "games_played": games_played,
            "weekly": weekly_stars,
            "skills": skills,
            "achievements": achievements,
            "time_spent": total_time_spent,
            "streak": streak,
            "played_today": list(played_today)
        })

    return jsonify({})



# ─── RUN ───────────────────────────────────

if __name__ == '__main__':
    app.run(debug=True)